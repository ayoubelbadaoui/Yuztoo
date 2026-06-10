"""Generate `docs/Yuztoo_Etat_Sprint.xlsx` — a client-facing status sheet.

Re-run after every sprint to regenerate a fresh file. Single source of truth
for the state communicated to the client; do not edit the .xlsx by hand.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------------------------------------------------------
# Data
# ----------------------------------------------------------------------------

HEADERS = [
    "#",
    "Sujet",
    "Statut",
    "Ce qui a été livré",
    "Comment vérifier",
]

# Status palette — keep in sync with the legend block at the bottom of the
# sheet so the client can decode the colours at a glance.
STATUS_DONE = "Corrigé"
STATUS_PARTIAL = "Partiel — chantier hors appli"
STATUS_BLOCKED_API = "Bloqué — clé API requise"
STATUS_PENDING_PRODUCT = "À clarifier (produit)"
STATUS_OTHER_OWNER = "Côté Ayoub"

ROWS: list[tuple[int, str, str, str, str]] = [
    (
        1,
        "Onboarding commerçant / client — appareil photo",
        STATUS_DONE,
        "L'accès à la caméra et le recadrage de la photo fonctionnent dans les deux parcours, avec messages clairs en français si une autorisation manque.",
        "Lancer l'inscription, prendre une photo de profil, vérifier que le recadrage s'ouvre puis que la photo s'enregistre.",
    ),
    (
        2,
        "Sous-catégories en fonction de la catégorie choisie",
        STATUS_DONE,
        "Une liste différente s'affiche selon la catégorie (Restaurant, Commerce de détail, Beauté & Bien-être, Sport & Fitness, Services). Un salon de beauté ne voit plus « Boulangerie » ou « Glacier ». La catégorie « Autre » passe directement à l'étape suivante.",
        "Onboarding pro → choisir successivement chaque catégorie et observer la grille de sous-catégories : elle change à chaque fois.",
    ),
    (
        3,
        "Saisie du téléphone — accepter le 0 devant les 9 chiffres",
        STATUS_DONE,
        "On peut taper son numéro avec ou sans le 0 initial, le système le formate proprement au format international.",
        "Taper « 06 12 34 56 78 » puis « 6 12 34 56 78 » → les deux sont acceptés et stockés à l'identique.",
    ),
    (
        4,
        "Date de naissance — design + double trait jaune",
        STATUS_DONE,
        "Sélecteur propre (style iOS), plus de double trait jaune, lecture correcte sur tous les écrans.",
        "Ouvrir le sélecteur de date sur un profil ou un onboarding → l'apparence est uniforme.",
    ),
    (
        5,
        "Bleu des champs de saisie — plus clair et homogène",
        STATUS_DONE,
        "Tous les champs (e-mail, téléphone, ville, code pays, etc.) partagent la même couleur de fond claire. Plus d'écart d'une page à l'autre.",
        "Comparer un champ d'inscription, un champ de modification de profil et un champ de notification : même teinte.",
    ),
    (
        6,
        "« Zone dangereuse » du profil — moins effrayante",
        STATUS_DONE,
        "La carte qui contient la suppression du compte n'est plus rouge. Seule la confirmation finale reste rouge — pour ne pas faire peur à un utilisateur qui ouvre son profil sans intention de supprimer.",
        "Ouvrir Profil pro → Réglages → la section est neutre. Tapper « Supprimer le compte » → la confirmation est bien rouge.",
    ),
    (
        7,
        "Validation de passage uniquement après scan QR / NFC",
        STATUS_DONE,
        "Impossible de marquer un passage sans avoir scanné.",
        "Tenter de valider un passage manuellement depuis le commerçant → bloqué.",
    ),
    (
        8,
        "Commerçants peuvent promouvoir d'autres commerces",
        STATUS_DONE,
        "Section partenaires fonctionnelle.",
        "Côté pro → Partenaires → inviter un autre commerce.",
    ),
    (
        9,
        "Heures d'ouverture — insensibles à la casse + pré-choix",
        STATUS_DONE,
        "Listes déroulantes propres, plus de saisie libre fragile.",
        "Modifier la vitrine → Horaires → choisir parmi les options.",
    ),
    (
        10,
        "Bouton « Modifier mon profil pro » opérant",
        STATUS_DONE,
        "Le bouton ouvre bien la modification du profil pro et plus la vitrine.",
        "Profil → Modifier mon profil pro → vous arrivez sur les infos de l'entreprise.",
    ),
    (
        11,
        "Ciblage des notifications (Habitués / Nouveaux / Tous)",
        STATUS_DONE,
        "Un client passé en « Habitué » ne reçoit plus une notif ciblée « Nouveaux », et inversement.",
        "Envoyer une notif « Nouveaux » → seuls les clients sans passages reçoivent.",
    ),
    (
        12,
        "Photo de profil ≠ logo",
        STATUS_DONE,
        "Les deux images sont distinctes et indépendantes.",
        "Mettre une photo de profil et un logo différents → les deux sont préservés.",
    ),
    (
        13,
        "Recherche commerce — approximative (Découvrir)",
        STATUS_DONE,
        "Tape une partie du nom (avec faute d'accent ou de casse) → trouve quand même.",
        "Découvrir → taper « boulang » → trouve « Boulangerie X ».",
    ),
    (
        14,
        "Compteur « 2 clients » à la première connexion",
        STATUS_DONE,
        "Le compte commerçant n'est plus compté comme un client.",
        "Inscrire un seul client → le compteur affiche 1.",
    ),
    (
        15,
        "Clavier figé pendant la saisie d'une notification",
        STATUS_DONE,
        "Le clavier reste fluide.",
        "Composer une notif longue → pas de blocage.",
    ),
    (
        16,
        "Photo de profil commerçant — chargement",
        STATUS_DONE,
        "La photo s'affiche immédiatement après upload.",
        "Changer la photo → elle apparaît sans relancer l'app.",
    ),
    (
        17,
        "Mots anglais lors de l'utilisation de la caméra",
        STATUS_DONE,
        "Tout est en français (Annuler / Confirmer / Recadrer).",
        "Prendre une photo dans n'importe quel onboarding.",
    ),
    (
        18,
        "Tap sur notification / promo → ouvre la promo",
        STATUS_DONE,
        "On arrive directement sur le détail de la promotion, plus sur la vitrine.",
        "Recevoir une notif promo → tapper → la fiche promo s'ouvre.",
    ),
    (
        19,
        "Bon disparaît après usage",
        STATUS_DONE,
        "Une fois utilisé, le bon n'est plus dans « Mes avantages ».",
        "Utiliser un bon → revenir → il a disparu.",
    ),
    (
        20,
        "Création compte pro depuis client — ne re-demande plus prénom / nom / DOB",
        STATUS_DONE,
        "Le système réutilise les infos déjà saisies. On ne demande que ce qui manque.",
        "Compte client → Créer un compte pro → le formulaire est court.",
    ),
    (
        21,
        "Notification ciblée — texte blanc sur blanc",
        STATUS_DONE,
        "Le texte saisi est lisible.",
        "Composer une notif → le texte est noir sur fond clair.",
    ),
    (
        22,
        "Carnet Yuztoo d'un pro en mode client (1ère bascule)",
        STATUS_DONE,
        "Le carnet se charge dès la première bascule, sans avoir besoin de suivre un commerce d'abord.",
        "Pro → bascule en mode client → le carnet est visible immédiatement.",
    ),
    (
        23,
        "Profil personnel → libellé « Créer un carnet Yuztoo » + auto-masquage",
        STATUS_DONE,
        "Le bon libellé s'affiche selon le type de compte. En plus : une fois que les deux comptes existent (client + pro), la proposition disparaît complètement.",
        "Profil pro → Informations personnelles → libellé correct. Une fois le carnet créé, la proposition disparaît.",
    ),
    (
        24,
        "Notif anniversaire — uniquement aux concernés",
        STATUS_DONE,
        "Seuls les clients dont c'est l'anniversaire reçoivent. Le système gère les cas complexes (changement de date, désabonnement-réabonnement, plusieurs modèles d'anniversaire actifs).",
        "Configurer une notif anniversaire et observer le journal. Seules les bonnes personnes reçoivent.",
    ),
    (
        25,
        "Notifications automatiques — opérantes + visibilité",
        STATUS_DONE,
        "Un panneau de diagnostic dans « Notifications automatiques » indique en temps réel si tout est en ordre, et un bouton « Tester l'envoi maintenant » permet une vérification immédiate.",
        "Côté pro → Notifications automatiques → panneau visible + bouton de test.",
    ),
    (
        26,
        "Tempo d'1 heure entre deux passages validés",
        STATUS_DONE,
        "Impossible d'enchaîner deux passages au même commerce dans l'heure. Le client voit un message clair s'il essaie.",
        "Valider un passage → réessayer immédiatement → message « patientez 1 heure ».",
    ),
    (
        27,
        "Vignettes actualités zoomables au tap",
        STATUS_DONE,
        "Tap sur une image d'actualité → ouverture plein écran avec zoom (pinch).",
        "Côté client → Actualités d'un commerce → tap sur image.",
    ),
    (
        28,
        "Comptage des vues de profil",
        STATUS_DONE,
        "Chaque visite unique d'un client sur la vitrine est comptabilisée. Le commerçant voit le total dans son tableau de bord.",
        "Côté pro → Statistiques → compteur de vues.",
    ),
    (
        29,
        "Bouton « Ne plus suivre » — discret, en bas",
        STATUS_DONE,
        "Le gros bouton du haut a été retiré. Un lien discret « Ne plus suivre » est désormais en bas du profil de la vitrine, après le contenu.",
        "Vitrine d'un commerce suivi → faire défiler tout en bas.",
    ),
    (
        30,
        "Notification reflète le nom courant du commerce",
        STATUS_DONE,
        "Si l'entreprise change de nom, toutes ses notifications futures (manuelles ET automatiques) utilisent le nouveau nom.",
        "Renommer le commerce → envoyer une notif → le nouveau nom apparaît.",
    ),
    (
        31,
        "Wizard fidélité côté client — apparition trop précoce",
        STATUS_DONE,
        "Le client ne voit le filtre par catégorie que s'il possède des cartes dans plusieurs catégories. Au début, il voit simplement tous ses commerces sans étape supplémentaire.",
        "Nouveau client qui ouvre la fidélité → vue simple, pas de wizard.",
    ),
    (
        32,
        "Vignette Yuztoo dans le carnet — réservée aux commerçants",
        STATUS_DONE,
        "Un client « pur » ne voit plus la vignette Yuztoo. Seuls les comptes pro qui utilisent l'appli en mode client la voient.",
        "Compte client pur → carnet → pas de vignette. Compte pro en mode client → carnet → vignette présente.",
    ),
    (
        33,
        "Bouton « Tout lire » alertes — format corrigé",
        STATUS_DONE,
        "Remplacé par une icône claire (double check), équilibre visuel avec l'icône poubelle à côté.",
        "Onglet alertes → coin haut droit.",
    ),
    (
        34,
        "Cadeau de bienvenue — section dédiée et visible",
        STATUS_DONE,
        "Une nouvelle section « Cadeaux de bienvenue » apparaît directement sur l'accueil client, au-dessus des promotions. Tap sur une carte → on bascule sur l'onglet fidélité pour récupérer le bon.",
        "Suivre un nouveau commerce qui a un cadeau de bienvenue → revenir à l'accueil → la section est visible immédiatement.",
    ),
    (
        35,
        "Premier scan = suivre + cadeau (pas validation de passage)",
        STATUS_DONE,
        "Au premier scan, on propose de suivre le commerce et on montre le cadeau de bienvenue. La validation de passage n'est plus proposée automatiquement — elle reste accessible depuis la vitrine pour les visites suivantes.",
        "Scanner pour la première fois un nouveau commerce → suivre → cadeau visible → on est posé sur la vitrine, pas dans une fenêtre de passage.",
    ),
    (
        36,
        "Email réutilisable après échec de la vérification mobile",
        STATUS_DONE,
        "Si la vérif SMS échoue ou s'interrompt, l'email et le numéro sont automatiquement libérés. L'utilisateur peut recommencer immédiatement avec les mêmes identifiants.",
        "Démarrer une inscription, abandonner / refaire → l'email est de nouveau accepté.",
    ),
    (
        37,
        "Différence visible entre statuts de clients (Habitué / Nouveau)",
        STATUS_DONE,
        "Le système les distingue désormais, et le ciblage de notifications respecte la distinction.",
        "Voir le ciblage à la création d'une notification.",
    ),
    (
        38,
        "« Modifier profil pro » arrive sur le formulaire pro",
        STATUS_DONE,
        "Le bouton ouvre directement la fiche pro à modifier (et non la vitrine).",
        "Profil → Modifier profil pro.",
    ),
    (
        39,
        "Tab bar pro — icône Tab 2 + bascule « nouveau passage » en Tab 1",
        STATUS_PENDING_PRODUCT,
        "Pas livré. Besoin de votre validation produit : quelle icône précise pour Tab 2 (rappels) et comment réorganiser Tab 1 sans perturber le flow commerçant existant.",
        "À définir ensemble.",
    ),
    (
        40,
        "Si pas Yuztoo installé → proposer l'installation",
        STATUS_PARTIAL,
        "Dans l'appli, le funnel est complet (invité → « Connectez-vous pour suivre »). Pour qu'un scan QR depuis l'appareil photo natif (sans Yuztoo installé) renvoie vers l'App Store / Play Store, il faut une page web yuztoo.app/vitrine/{id} + activer Universal Links iOS et App Links Android. Chantier infrastructure (hébergement web), pas un chantier code Flutter.",
        "À planifier avec votre équipe DevOps / hébergement.",
    ),
    (
        41,
        "Champ adresse — autocomplétion Google Places",
        STATUS_BLOCKED_API,
        "Pas livré. L'intégration nécessite une clé API Google Maps activée + facturation côté Google Cloud. Sans ça, l'autocomplete ne peut pas tourner.",
        "À débloquer côté Google Cloud puis on intègre.",
    ),
    (
        42,
        "Scan NFC sur les plaques (nouvelle version)",
        STATUS_OTHER_OWNER,
        "Architecture en place, implémentation de la nouvelle version portée par Ayoub.",
        "À voir avec lui.",
    ),
]


# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------

NAVY = "0E2A44"        # header background — Yuztoo brand navy
WHITE = "FFFFFF"
GREEN = "C6EFCE"
GREEN_TXT = "006100"
ORANGE = "FFEB9C"
ORANGE_TXT = "9C5700"
RED = "FFC7CE"
RED_TXT = "9C0006"
GREY = "D9D9D9"
GREY_TXT = "595959"

STATUS_FILL = {
    STATUS_DONE: (GREEN, GREEN_TXT),
    STATUS_PARTIAL: (ORANGE, ORANGE_TXT),
    STATUS_BLOCKED_API: (RED, RED_TXT),
    STATUS_PENDING_PRODUCT: (RED, RED_TXT),
    STATUS_OTHER_OWNER: (GREY, GREY_TXT),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def write_header(ws: Worksheet) -> None:
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def write_rows(ws: Worksheet) -> None:
    body_font = Font(size=11, name="Calibri")
    for row in ROWS:
        ws.append(row)
        excel_row = ws.max_row
        status_value = row[2]
        fill_color, font_color = STATUS_FILL[status_value]
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="left",
            )
            cell.border = THIN_BORDER
        # Highlight the status column with the matching colour and recolour
        # the text so it stays readable on the coloured background.
        status_cell = ws.cell(row=excel_row, column=3)
        status_cell.fill = PatternFill("solid", fgColor=fill_color)
        status_cell.font = Font(
            size=11, name="Calibri", color=font_color, bold=True
        )
        status_cell.alignment = Alignment(
            vertical="center", horizontal="center", wrap_text=True
        )


def auto_size(ws: Worksheet) -> None:
    # Column widths tuned for readability — auto-sizing on wrapped text is
    # unreliable with openpyxl, so we set explicit widths.
    widths = {
        1: 5,    # #
        2: 38,   # Sujet
        3: 22,   # Statut
        4: 60,   # Ce qui a été livré
        5: 50,   # Comment vérifier
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row heights: bumped a touch so multi-line cells breathe.
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60


def write_summary(ws: Worksheet) -> None:
    """Append a small summary block + colour legend below the data."""
    ws.append([])  # blank line
    ws.append([])

    counts: dict[str, int] = {}
    for r in ROWS:
        counts[r[2]] = counts.get(r[2], 0) + 1

    title_row = ws.max_row + 1
    ws.cell(row=title_row, column=1, value="Récapitulatif").font = Font(
        bold=True, size=12, name="Calibri"
    )

    legend = [
        ("✅ Corrigé et livré", STATUS_DONE),
        ("⚠️ Partiel — chantier hors appli", STATUS_PARTIAL),
        ("⏸ Bloqué — clé API requise", STATUS_BLOCKED_API),
        ("⏸ À clarifier (produit)", STATUS_PENDING_PRODUCT),
        ("🔧 Côté Ayoub", STATUS_OTHER_OWNER),
    ]
    for label, status_key in legend:
        ws.append([label, "", str(counts.get(status_key, 0))])
        row = ws.max_row
        fill_color, font_color = STATUS_FILL[status_key]
        ws.cell(row=row, column=3).fill = PatternFill(
            "solid", fgColor=fill_color
        )
        ws.cell(row=row, column=3).font = Font(
            size=11, name="Calibri", bold=True, color=font_color
        )
        ws.cell(row=row, column=3).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[row].height = 22


def main() -> None:
    output = Path(__file__).resolve().parent.parent / "docs" / "Yuztoo_Etat_Sprint.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "État sprint"

    write_header(ws)
    write_rows(ws)
    auto_size(ws)
    write_summary(ws)

    # Freeze the header so it stays visible while scrolling — saves the
    # client from losing the column meaning halfway down the sheet.
    ws.freeze_panes = "A2"

    wb.save(output)
    print(f"wrote {output}")


if __name__ == "__main__":
    main()
